"""Inventory file parsers - CSV and XLSX.

Master Spec §15 Phase 3: "Capability list ingest (Excel upload + AI
extraction with redaction)". This module turns the raw artifact bytes
into a row-shaped representation the LLM can reason about. The LLM does
the column-mapping; we just give it well-shaped rows.

Phase 3 only supports CSV + XLSX. PDF ingest is a Phase 6 hardening
target (table extraction is a different problem and the inventory
documents Eugene's customers ship are reliably tabular).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from typing import Any


class UnsupportedInventoryFormat(ValueError):
    """Raised when an artifact's MIME isn't a recognized inventory format."""


# MIME types that the ingest endpoint accepts.
SUPPORTED_MIME = {
    "text/csv": "csv",
    "text/plain": "csv",  # treat .txt as CSV; most inventory exports save this way
    "application/vnd.ms-excel": "xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
}

# Max rows we ship to the LLM. Above this and we'd either bust the context
# window or pay a fortune in tokens. v1 inventories are typically 50-300 rows.
MAX_ROWS = 500


def kind_for_mime(mime_type: str) -> str:
    try:
        return SUPPORTED_MIME[mime_type]
    except KeyError as exc:
        raise UnsupportedInventoryFormat(
            f"Inventory format {mime_type!r} not supported. Accept CSV or XLSX."
        ) from exc


def parse_inventory(data: bytes, mime_type: str) -> list[dict[str, Any]]:
    """Parse `data` into a list of row-dicts. Header row becomes the keys.

    Returns at most MAX_ROWS rows; the last row in the response is a
    sentinel `{"__truncated__": True}` marker when the input was longer.
    """
    kind = kind_for_mime(mime_type)
    if kind == "csv":
        rows = _parse_csv(data)
    elif kind == "xlsx":
        rows = _parse_xlsx(data)
    else:
        raise UnsupportedInventoryFormat(f"Unknown internal kind {kind!r}.")

    out = list(rows)
    truncated = len(out) > MAX_ROWS
    if truncated:
        out = out[:MAX_ROWS]
        out.append({"__truncated__": True, "__hint__": f"Input had > {MAX_ROWS} rows."})
    return out


def _parse_csv(data: bytes) -> Iterable[dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        # Strip whitespace from keys + values for stability.
        yield {(k or "").strip(): (v or "").strip() for k, v in row.items() if k is not None}


def _parse_xlsx(data: bytes) -> Iterable[dict[str, Any]]:
    # openpyxl is lazy-imported so test runs that don't touch XLSX don't
    # pay the import cost.
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)]
    for raw in rows_iter:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        yield {
            headers[i]: ("" if v is None else str(v).strip())
            for i, v in enumerate(raw)
            if i < len(headers)
        }
