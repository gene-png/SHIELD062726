"""Deliverable renderers - turn a capability list into XLSX + PDF bytes.

Master Spec §15 Phase 3: "PDF + XLSX exporters for the deliverable."

XLSX: openpyxl. Header row + one row per capability + a summary row at
the bottom (Total Cost, Estimated Savings).

PDF: ReportLab. Pure Python; no native deps required (unlike WeasyPrint).
Phase 6 polish can revisit visual fidelity, but for v1 the deliverable is
a real, legitimate PDF with a title, summary, table, and savings figure.

Both renderers are pure functions over the data; no DB, no I/O. The
route layer writes the bytes via the existing StorageBackend.
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass

from app.models.capability import CapabilityDisposition, CapabilityItem, CapabilityList


@dataclass(frozen=True)
class DeliverableContext:
    """Inputs the renderers share. Built once by the route layer."""

    client_legal_name: str
    service_title: str
    cap_list: CapabilityList
    items: list[CapabilityItem]
    total_cost: float
    estimated_savings: float
    savings_cost_known: bool


def _disposition_label(d: CapabilityDisposition | None) -> str:
    if d is None:
        return "Undecided"
    return {
        CapabilityDisposition.KEEP: "Keep",
        CapabilityDisposition.CONSOLIDATE: "Consolidate",
        CapabilityDisposition.CUT: "Cut",
    }[d]


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    cap_list: CapabilityList,
    items: Iterable[CapabilityItem],
) -> DeliverableContext:
    items_list = list(items)
    total_cost = 0.0
    estimated_savings = 0.0
    savings_known = True
    for it in items_list:
        if it.annual_cost_usd is not None:
            total_cost += float(it.annual_cost_usd)
        if it.disposition == CapabilityDisposition.CUT:
            if it.annual_cost_usd is None:
                savings_known = False
            else:
                estimated_savings += float(it.annual_cost_usd)
    return DeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        cap_list=cap_list,
        items=items_list,
        total_cost=total_cost,
        estimated_savings=estimated_savings,
        savings_cost_known=savings_known,
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: DeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned no active worksheet")
    ws.title = "Capability List"

    header = [
        "Name",
        "Vendor",
        "Category",
        "Function",
        "Annual Cost (USD)",
        "Licenses",
        "Disposition",
        "Rationale",
        "Notes",
        "AI Confidence %",
    ]
    ws.append(header)
    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for item in ctx.items:
        ws.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                item.function or "",
                float(item.annual_cost_usd) if item.annual_cost_usd is not None else "",
                item.license_count if item.license_count is not None else "",
                _disposition_label(item.disposition),
                item.disposition_rationale or "",
                item.notes or "",
                item.confidence_pct if item.confidence_pct is not None else "",
            ]
        )

    # Summary row at the bottom.
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Total annual cost").font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=ctx.total_cost).number_format = "$#,##0"
    ws.cell(row=summary_row + 1, column=1, value="Estimated annual savings").font = Font(bold=True)
    savings_cell = ws.cell(row=summary_row + 1, column=5, value=ctx.estimated_savings)
    savings_cell.number_format = "$#,##0"
    if not ctx.savings_cost_known:
        ws.cell(
            row=summary_row + 1,
            column=6,
            value="≥ (one or more cut rows missing a cost)",
        ).font = Font(italic=True)

    # Reasonable column widths.
    widths = [28, 22, 16, 28, 18, 10, 14, 38, 38, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_pdf(ctx: DeliverableContext) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(ctx.client_legal_name, body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Summary", h2))
    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    story.append(
        Paragraph(
            f"Capabilities reviewed: <b>{len(ctx.items)}</b> · "
            f"Total annual cost: <b>${ctx.total_cost:,.0f}</b> · "
            f"Estimated annual savings: <b>{savings}</b>",
            body,
        )
    )
    if not ctx.savings_cost_known:
        story.append(
            Paragraph(
                "Note: at least one row marked <i>Cut</i> is missing an annual cost. "
                "The savings figure is a lower bound.",
                body,
            )
        )

    story.append(Paragraph("Capability list", h2))

    table_data: list[list] = [["Name", "Vendor", "Category", "Annual cost", "Disposition"]]
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        table_data.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )

    table = Table(
        table_data,
        colWidths=[2.2 * inch, 1.4 * inch, 1.2 * inch, 1.0 * inch, 1.2 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    return out.getvalue()


# ---------------------------------------------------------------------------
# DOCX (Work Order C4) - mirrors the PDF content.
# ---------------------------------------------------------------------------


def render_docx(ctx: DeliverableContext) -> bytes:
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    add_heading(doc, "Summary")
    lines = [
        f"Capabilities reviewed: {len(ctx.items)}",
        f"Total annual cost: ${ctx.total_cost:,.0f}",
        f"Estimated annual savings: {savings}",
    ]
    if not ctx.savings_cost_known:
        lines.append(
            "Note: at least one row marked Cut is missing an annual cost. "
            "The savings figure is a lower bound."
        )
    add_paragraphs(doc, lines)

    add_heading(doc, "Capability list")
    rows = []
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        rows.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )
    add_table(doc, ["Name", "Vendor", "Category", "Annual cost", "Disposition"], rows)

    return to_bytes(doc)
