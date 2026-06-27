"""Zero Trust deliverable renderers - PDF + XLSX from a ZT assessment.

Three-sheet XLSX (Score Summary / Answers / Gap Plan) + executive PDF
(overall stage + per-pillar table, then top-N gap table). Pure
functions, no I/O. Framework-aware: CISA labels render Traditional/
Initial/..., DoD labels render Baseline/Target/...
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass
from typing import TYPE_CHECKING

from app.models.zt_assessment import ZtAnswer, ZtAssessment
from app.zt.catalog import capabilities, pillars
from app.zt.maturity import ZtFrameworkCode, stage_label
from app.zt.scoring import GapAnalysis, ScoreResult

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle


@dataclass(frozen=True)
class ZtDeliverableContext:
    client_legal_name: str
    service_title: str
    framework: ZtFrameworkCode
    assessment: ZtAssessment
    answers: list[ZtAnswer]
    score: ScoreResult
    gap: GapAnalysis


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    framework: ZtFrameworkCode,
    assessment: ZtAssessment,
    answers: Iterable[ZtAnswer],
    score: ScoreResult,
    gap: GapAnalysis,
) -> ZtDeliverableContext:
    return ZtDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        framework=framework,
        assessment=assessment,
        answers=list(answers),
        score=score,
        gap=gap,
    )


def _fmt(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}"


def _framework_label(framework: ZtFrameworkCode) -> str:
    return (
        "CISA ZTMM 2.0"
        if framework == ZtFrameworkCode.CISA_ZTMM_2_0
        else "DoD ZT Reference Architecture"
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: ZtDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Score Summary ---
    ws = wb.create_sheet("Score Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Framework", _framework_label(ctx.framework)])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Overall stage", ctx.score.overall_stage_label])
    ws.append(["Average stage", _fmt(ctx.score.average_stage)])
    ws.append(["Coverage", f"{ctx.score.answered_capabilities}/{ctx.score.total_capabilities}"])
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    ws.append(["Pillar", "Name", "Answered", "Total", "Coverage %", "Average stage"])
    for col_idx in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
    for ps in ctx.score.by_pillar:
        ws.append(
            [
                ps.pillar_code,
                ps.pillar_name,
                ps.answered_count,
                ps.capability_count,
                ps.coverage_pct,
                _fmt(ps.average_stage),
            ]
        )
    for w, col in zip([10, 36, 12, 10, 14, 16], range(1, 7), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Answers ---
    ws2 = wb.create_sheet("Answers")
    headers = ["Capability", "Pillar", "Name", "Outcome", "Stage", "Stage label", "Notes"]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    answers_by_code = {a.capability_code: a for a in ctx.answers}
    # Use catalog order so missing answers still render as blank rows.
    pillar_lookup = {p.code: p.name for p in pillars(ctx.framework)}
    for cap in capabilities(ctx.framework):
        ans = answers_by_code.get(cap.code)
        s = ans.maturity_stage if ans else None
        notes = ans.notes if ans else None
        ws2.append(
            [
                cap.code,
                f"{cap.pillar_code} · {pillar_lookup.get(cap.pillar_code, cap.pillar_code)}",
                cap.name,
                cap.outcome,
                s if s is not None else "",
                stage_label(s, ctx.framework) if s is not None else "Unscored",
                notes or "",
            ]
        )
    for w, col in zip([18, 30, 36, 60, 8, 16, 60], range(1, 8), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Gap Plan ---
    ws3 = wb.create_sheet("Gap Plan")
    headers3 = [
        "Capability",
        "Pillar",
        "Name",
        "Current stage",
        "Target stage",
        "Gap size",
        "Priority",
        "Notes",
    ]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    for g in ctx.gap.gaps:
        ws3.append(
            [
                g.code,
                g.pillar_code,
                g.name,
                g.current_stage,
                g.target_stage,
                g.gap_size,
                g.priority_score,
                g.notes or "",
            ]
        )
    if not ctx.gap.gaps:
        ws3.append(["—", "", "No gaps at target stage", "", ctx.gap.target_stage, 0, 0, ""])
        ws3.cell(row=2, column=3).font = italic
    for w, col in zip([18, 10, 36, 14, 14, 12, 12, 50], range(1, 9), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_docx(ctx: ZtDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(
        doc,
        ctx.service_title,
        f"{ctx.client_legal_name} · {_framework_label(ctx.framework)}",
    )

    add_heading(doc, "Maturity summary")
    add_paragraphs(
        doc,
        [
            f"Overall stage: {ctx.score.overall_stage_label}",
            f"Average stage: {_fmt(ctx.score.average_stage)}",
            f"Coverage: {ctx.score.answered_capabilities}/"
            f"{ctx.score.total_capabilities} ({ctx.score.coverage_pct}%)",
        ],
    )

    add_heading(doc, "Per-pillar rollup")
    add_table(
        doc,
        ["Pillar", "Name", "Average stage", "Coverage"],
        [
            [
                ps.pillar_code,
                ps.pillar_name,
                _fmt(ps.average_stage),
                f"{ps.answered_count}/{ps.capability_count} ({ps.coverage_pct}%)",
            ]
            for ps in ctx.score.by_pillar
        ],
    )

    add_heading(doc, f"Top remediation gaps (target S{ctx.gap.target_stage})")
    if not ctx.gap.gaps:
        add_paragraphs(
            doc,
            [f"No gaps at target stage {ctx.gap.target_stage} " f"({ctx.gap.target_label})."],
        )
    else:
        add_table(
            doc,
            ["Code", "Pillar", "Capability", "Current → Target", "Priority"],
            [
                [
                    g.code,
                    g.pillar_code,
                    g.name,
                    f"S{g.current_stage} → S{g.target_stage}",
                    f"{g.priority_score:.2f}",
                ]
                for g in ctx.gap.gaps
            ],
        )

    return to_bytes(doc)


def render_pdf(ctx: ZtDeliverableContext) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(f"{ctx.client_legal_name} · {_framework_label(ctx.framework)}", body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Maturity summary", h2))
    story.append(
        Paragraph(
            f"Overall stage: <b>{ctx.score.overall_stage_label}</b> · "
            f"Average stage: <b>{_fmt(ctx.score.average_stage)}</b> · "
            f"Coverage: <b>{ctx.score.answered_capabilities}/"
            f"{ctx.score.total_capabilities}</b> "
            f"({ctx.score.coverage_pct}%)",
            body,
        )
    )

    story.append(Paragraph("Per-pillar rollup", h2))
    fn_table_data: list[list] = [["Pillar", "Name", "Average stage", "Coverage"]]
    for ps in ctx.score.by_pillar:
        fn_table_data.append(
            [
                ps.pillar_code,
                ps.pillar_name,
                _fmt(ps.average_stage),
                f"{ps.answered_count}/{ps.capability_count} ({ps.coverage_pct}%)",
            ]
        )
    fn_table = Table(
        fn_table_data,
        colWidths=[0.8 * inch, 3.2 * inch, 1.2 * inch, 1.6 * inch],
        repeatRows=1,
    )
    fn_table.setStyle(_table_style())
    story.append(fn_table)

    story.append(PageBreak())

    story.append(Paragraph(f"Top remediation gaps (target S{ctx.gap.target_stage})", h2))
    if not ctx.gap.gaps:
        story.append(
            Paragraph(
                f"No gaps at target stage {ctx.gap.target_stage} " f"({ctx.gap.target_label}).",
                body,
            )
        )
    else:
        gap_table_data: list[list] = [
            ["Code", "Pillar", "Capability", "Current → Target", "Priority"]
        ]
        for g in ctx.gap.gaps:
            gap_table_data.append(
                [
                    g.code,
                    g.pillar_code,
                    g.name,
                    f"S{g.current_stage} → S{g.target_stage}",
                    f"{g.priority_score:.2f}",
                ]
            )
        gap_table = Table(
            gap_table_data,
            colWidths=[1.1 * inch, 0.8 * inch, 2.9 * inch, 1.4 * inch, 0.8 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = [
    "ZtDeliverableContext",
    "build_context",
    "render_pdf",
    "render_xlsx",
]
