"""ATT&CK PDF + XLSX exporter smokes."""

from __future__ import annotations

import io
import uuid

import pytest
from app.attack.analytics import compute as compute_heatmap
from app.attack.catalog import TECHNIQUES
from app.attack.coverage import CoverageStatus
from app.attack.exporters import build_context, render_pdf, render_xlsx
from app.models.attack_assessment import (
    AttackAssessment,
    AttackAssessmentStatus,
    AttackCoverage,
)


def _build_inputs(*, default_status: str | None = "covered"):
    a = AttackAssessment(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=AttackAssessmentStatus.APPROVED,
    )
    coverage: list[AttackCoverage] = []
    for t in TECHNIQUES:
        coverage.append(
            AttackCoverage(
                id=uuid.uuid4(),
                assessment_id=a.id,
                technique_code=t.id,
                status=default_status,
            )
        )
    coverage_map = {c.technique_code: c.status for c in coverage}
    rollup = compute_heatmap(coverage_map)
    return a, coverage, rollup


def _ctx(*, default_status: str | None = "covered"):
    a, coverage, rollup = _build_inputs(default_status=default_status)
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="MITRE ATT&CK Coverage",
        assessment=a,
        coverage=coverage,
        rollup=rollup,
    )


@pytest.mark.unit
def test_xlsx_has_three_sheets() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx())
    assert raw[:2] == b"PK"
    wb = load_workbook(io.BytesIO(raw))
    assert set(wb.sheetnames) == {"Heatmap Summary", "Coverage", "Gaps"}


@pytest.mark.unit
def test_xlsx_coverage_sheet_has_one_row_per_technique() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx())
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Coverage"]
    assert ws.max_row == len(TECHNIQUES) + 1


@pytest.mark.unit
def test_xlsx_gap_sheet_lists_only_gaps() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(default_status=CoverageStatus.GAP.value)
    raw = render_xlsx(ctx)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Gaps"]
    # Header + every technique.
    assert ws.max_row == len(TECHNIQUES) + 1


@pytest.mark.unit
def test_xlsx_gap_placeholder_when_no_gaps() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(default_status=CoverageStatus.COVERED.value)
    raw = render_xlsx(ctx)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Gaps"]
    assert ws.max_row == 2  # header + single placeholder
    assert ws.cell(row=2, column=2).value == "No gaps recorded"


@pytest.mark.unit
def test_pdf_renders_valid_bytes() -> None:
    raw = render_pdf(_ctx())
    assert raw.startswith(b"%PDF-")
    assert len(raw) > 2000


@pytest.mark.unit
def test_pdf_handles_zero_gaps() -> None:
    ctx = _ctx(default_status=CoverageStatus.COVERED.value)
    raw = render_pdf(ctx)
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_build_context_falls_back_when_client_none() -> None:
    a, coverage, rollup = _build_inputs()
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        assessment=a,
        coverage=coverage,
        rollup=rollup,
    )
    assert ctx.client_legal_name == "Client"
