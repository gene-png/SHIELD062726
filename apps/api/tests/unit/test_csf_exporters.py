"""Smoke tests for the CSF PDF + XLSX exporters."""

from __future__ import annotations

import io
import uuid

import pytest
from app.csf.catalog import SUBCATEGORIES
from app.csf.exporters import build_context, render_pdf, render_xlsx
from app.csf.gap import analyze as analyze_gaps
from app.csf.scoring import compute as compute_score
from app.models.csf_assessment import CsfAnswer, CsfAssessment, CsfAssessmentStatus


def _build_inputs(*, answers_tier: int | None = 3) -> tuple[CsfAssessment, list[CsfAnswer]]:
    a = CsfAssessment(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=CsfAssessmentStatus.APPROVED,
    )
    answers = []
    for sc in SUBCATEGORIES:
        ans = CsfAnswer(
            id=uuid.uuid4(),
            assessment_id=a.id,
            subcategory_code=sc.code,
            maturity_tier=answers_tier,
        )
        answers.append(ans)
    return a, answers


@pytest.fixture()
def context_with_full_tier3():
    a, answers = _build_inputs(answers_tier=3)
    tier_map = {ans.subcategory_code: ans.maturity_tier for ans in answers}
    score = compute_score(tier_map)
    gap = analyze_gaps(tier_map, target_tier=4)  # everything is below T4
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="NIST CSF 2.0 Assessment",
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )


@pytest.mark.unit
def test_xlsx_render_has_three_sheets(context_with_full_tier3) -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(context_with_full_tier3)
    assert isinstance(raw, bytes)
    assert raw[:2] == b"PK"  # XLSX is a zip envelope
    wb = load_workbook(io.BytesIO(raw))
    assert set(wb.sheetnames) == {"Score Summary", "Answers", "Gap Plan"}


@pytest.mark.unit
def test_xlsx_score_summary_carries_overall_label(context_with_full_tier3) -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(context_with_full_tier3)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Score Summary"]
    cells = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value) for r in range(1, 7)]
    label = dict(cells).get("Overall maturity")
    # All tier 3 -> Repeatable.
    assert label == "Repeatable"


@pytest.mark.unit
def test_xlsx_answers_sheet_has_one_row_per_subcategory(context_with_full_tier3) -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(context_with_full_tier3)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Answers"]
    # 106 subcategories + 1 header = 107 rows.
    assert ws.max_row == 107


@pytest.mark.unit
def test_xlsx_gap_plan_contains_top_gaps_when_target_is_adaptive(
    context_with_full_tier3,
) -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(context_with_full_tier3)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Gap Plan"]
    # Default top_n=20 with target=4 produces 20 gap rows + 1 header.
    assert ws.max_row == 21


@pytest.mark.unit
def test_pdf_render_produces_valid_pdf(context_with_full_tier3) -> None:
    raw = render_pdf(context_with_full_tier3)
    assert isinstance(raw, bytes)
    assert raw.startswith(b"%PDF-")
    assert len(raw) > 2000


@pytest.mark.unit
def test_pdf_handles_empty_gap_list() -> None:
    a, answers = _build_inputs(answers_tier=4)  # everyone Adaptive
    tier_map = {ans.subcategory_code: ans.maturity_tier for ans in answers}
    score = compute_score(tier_map)
    # Target tier 3 means everyone is past it -> zero gaps.
    gap = analyze_gaps(tier_map, target_tier=3)
    assert gap.total_gap_count == 0
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )
    raw = render_pdf(ctx)
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_xlsx_handles_empty_gap_list_with_placeholder_row() -> None:
    from openpyxl import load_workbook

    a, answers = _build_inputs(answers_tier=4)
    tier_map = {ans.subcategory_code: ans.maturity_tier for ans in answers}
    score = compute_score(tier_map)
    gap = analyze_gaps(tier_map, target_tier=3)
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    ws = wb["Gap Plan"]
    # Header + 1 placeholder row.
    assert ws.max_row == 2
    assert ws.cell(row=2, column=4).value == "No gaps at target tier"


@pytest.mark.unit
def test_build_context_falls_back_to_client_when_legal_name_is_none() -> None:
    a, answers = _build_inputs()
    tier_map = {ans.subcategory_code: ans.maturity_tier for ans in answers}
    score = compute_score(tier_map)
    gap = analyze_gaps(tier_map)
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )
    assert ctx.client_legal_name == "Client"
