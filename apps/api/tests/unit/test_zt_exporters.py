"""ZT PDF + XLSX exporter smokes for both frameworks."""

from __future__ import annotations

import io
import uuid

import pytest
from app.models.zt_assessment import (
    ZtAnswer,
    ZtAssessment,
    ZtAssessmentStatus,
    ZtFramework,
)
from app.zt.catalog import capabilities
from app.zt.exporters import build_context, render_pdf, render_xlsx
from app.zt.maturity import ZtFrameworkCode
from app.zt.scoring import analyze_gaps, compute


def _build_inputs(
    framework: ZtFrameworkCode, *, stage: int | None = 3
) -> tuple[ZtAssessment, list[ZtAnswer]]:
    db_framework = (
        ZtFramework.CISA_ZTMM_2_0
        if framework == ZtFrameworkCode.CISA_ZTMM_2_0
        else ZtFramework.DOD_ZTRA
    )
    a = ZtAssessment(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        framework=db_framework,
        version=1,
        status=ZtAssessmentStatus.APPROVED,
    )
    answers: list[ZtAnswer] = []
    for cap in capabilities(framework):
        answers.append(
            ZtAnswer(
                id=uuid.uuid4(),
                assessment_id=a.id,
                capability_code=cap.code,
                maturity_stage=stage,
            )
        )
    return a, answers


def _ctx(framework: ZtFrameworkCode, stage: int | None = 3, *, target: int = 4):
    a, answers = _build_inputs(framework, stage=stage)
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    score = compute(framework, stage_map)
    gap = analyze_gaps(framework, stage_map, target_stage=target)
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="Zero Trust Assessment",
        framework=framework,
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )


@pytest.mark.unit
def test_cisa_xlsx_has_three_sheets() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0))
    assert raw[:2] == b"PK"
    wb = load_workbook(io.BytesIO(raw))
    assert set(wb.sheetnames) == {"Score Summary", "Answers", "Gap Plan"}


@pytest.mark.unit
def test_dod_xlsx_renders() -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(_ctx(ZtFrameworkCode.DOD_ZTRA))
    wb = load_workbook(io.BytesIO(raw))
    assert "Score Summary" in wb.sheetnames
    ws = wb["Score Summary"]
    # Spot-check the framework cell.
    rows = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value) for r in range(1, 8)]
    fw = dict(rows).get("Framework")
    assert fw == "DoD ZT Reference Architecture"


@pytest.mark.unit
def test_cisa_xlsx_answers_sheet_row_count() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0))))
    ws = wb["Answers"]
    # 37 capabilities + 1 header.
    assert ws.max_row == 38


@pytest.mark.unit
def test_dod_xlsx_answers_sheet_row_count() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(render_xlsx(_ctx(ZtFrameworkCode.DOD_ZTRA))))
    ws = wb["Answers"]
    # 50 capabilities + 1 header.
    assert ws.max_row == 51


@pytest.mark.unit
def test_cisa_pdf_renders() -> None:
    raw = render_pdf(_ctx(ZtFrameworkCode.CISA_ZTMM_2_0))
    assert raw.startswith(b"%PDF-")
    assert len(raw) > 2000


@pytest.mark.unit
def test_dod_pdf_renders() -> None:
    raw = render_pdf(_ctx(ZtFrameworkCode.DOD_ZTRA))
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_pdf_handles_empty_gap_list() -> None:
    # Score everyone Optimal (stage 4) and target 3 -> zero gaps.
    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=4, target=3)
    raw = render_pdf(ctx)
    assert raw.startswith(b"%PDF-")


@pytest.mark.unit
def test_xlsx_handles_empty_gap_list_with_placeholder() -> None:
    from openpyxl import load_workbook

    ctx = _ctx(ZtFrameworkCode.CISA_ZTMM_2_0, stage=4, target=3)
    wb = load_workbook(io.BytesIO(render_xlsx(ctx)))
    ws = wb["Gap Plan"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=3).value == "No gaps at target stage"


@pytest.mark.unit
def test_build_context_falls_back_when_client_none() -> None:
    a, answers = _build_inputs(ZtFrameworkCode.CISA_ZTMM_2_0)
    stage_map = {ans.capability_code: ans.maturity_stage for ans in answers}
    score = compute(ZtFrameworkCode.CISA_ZTMM_2_0, stage_map)
    gap = analyze_gaps(ZtFrameworkCode.CISA_ZTMM_2_0, stage_map)
    ctx = build_context(
        client_legal_name=None,
        service_title="x",
        framework=ZtFrameworkCode.CISA_ZTMM_2_0,
        assessment=a,
        answers=answers,
        score=score,
        gap=gap,
    )
    assert ctx.client_legal_name == "Client"
